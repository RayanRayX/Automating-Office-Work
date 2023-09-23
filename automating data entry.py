import openpyxl
import os

def add_data_to_existing_excel():
    try:
        # List Excel files in the current directory
        excel_files = [f for f in os.listdir() if f.endswith(".xlsx")]

        if not excel_files:
            print("No Excel files (.xlsx) found in the current directory.")
            return

        # Display available Excel files
        print("Available Excel files in the current directory:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")

        # Ask the user to select an existing Excel file
        file_index = int(input("Enter the number corresponding to the Excel file where you want to add data: ")) - 1

        # Get the selected file name
        file_name = excel_files[file_index]

        # Construct the full file path
        file_path = os.path.join(os.getcwd(), file_name)

        # Open the selected Excel file
        workbook = openpyxl.load_workbook(file_path)

        # Get a list of sheet names in the workbook
        sheet_names = workbook.sheetnames

        # Ask the user to select a sheet
        print("Available sheets in the Excel file:")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"{i}. {sheet_name}")

        sheet_index = int(input("Enter the number corresponding to the sheet where you want to add data: ")) - 1
        selected_sheet = workbook[sheet_names[sheet_index]]

        # Get the headers in the selected sheet
        headers = [cell.value for cell in selected_sheet[1]]

        # Find the row number for the first header (assuming you want to add data to the same row)
        row_number = len(selected_sheet["A"]) + 1

        while True:
            for header in headers:
                data = input(f"Enter data to add to '{header}': ").strip()

                # Find the column number for the selected header
                column_number = None
                for cell in selected_sheet[1]:
                    if cell.value == header:
                        column_number = cell.column
                        break

                if column_number is not None:
                    # Add data to the specified row and column for the header
                    cell = selected_sheet.cell(row=row_number, column=column_number)
                    cell.value = data

            # Save the changes to the Excel file
            workbook.save(file_path)

            another_set = input("Do you want to enter another set of content in the next column? (yes/no): ").strip().lower()
            if another_set != "yes":
                break

            # Move to the next row for the next set of content
            row_number += 1

        # Close the Excel file
        workbook.close()

        print("Data added successfully in '{file_name}'.")
    except (ValueError, IndexError):
        print("Invalid input. Please enter valid numbers for file and sheet selection.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage:
add_data_to_existing_excel()
